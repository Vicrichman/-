"""
喜过分析模板 - 数据提取脚本
从 喜过数据源收集表.xlsx 提取卡西欧 vs 蔻驰的全量运营数据
输出: data.js (内联到看板HTML)

核心修复:
1. 大盘指数提取覆盖全时间范围（不只是May）
2. 社区投放/得物推按月汇总
3. UV月度数据
4. SPU级别的UV、支付订单、支付金额详情
"""
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict
import json
import sys, os

# ── B patch: --output-dir support ──
_OUTPUT_DIR = None
for _i, _a in enumerate(sys.argv):
    if _a == '--output-dir' and _i + 1 < len(sys.argv):
        _OUTPUT_DIR = sys.argv[_i + 1]
        assert os.path.isdir(_OUTPUT_DIR), f"output-dir not found: {_OUTPUT_DIR}"
        break

EXCEL_PATH = "/mnt/e/Obsidian本地仓库/09-数据源/喜过数据源收集表.xlsx"
START_DATE = "2025-05-01"
END_DATE = None  # Dynamic: auto-detected from latest transaction data

def standardize_brand(raw):
    """Dynamic brand normalization — no whitelist."""
    if raw is None or (isinstance(raw, float) and raw != raw):
        return "未分类品牌"
    s = str(raw).strip()
    if s in ("", "-", "/", "暂无"):
        return "未分类品牌"
    sup = s.upper()
    if "CASIO" in sup or "卡西欧" in s: return "CASIO/卡西欧"
    if "COACH" in sup or "蔻驰" in s: return "COACH/蔻驰"
    if "SWATCH" in sup or "斯沃琪" in s: return "SWATCH/斯沃琪"
    if "GIVENCHY" in sup or "纪梵希" in s: return "Givenchy/纪梵希"
    return s

def brand_short(std_brand):
    mapping = {"CASIO/卡西欧":"卡西欧","COACH/蔻驰":"蔻驰","SWATCH/斯沃琪":"斯沃琪","Givenchy/纪梵希":"纪梵希","未分类品牌":"未分类品牌"}
    return mapping.get(std_brand, std_brand)

def excel_to_date(serial):
    """Convert Excel date serial or string to YYYY-MM-DD string"""
    if isinstance(serial, datetime):
        return serial.strftime("%Y-%m-%d")
    if isinstance(serial, str) and serial.strip():
        # Try to parse string dates like "2026-05-07 14:21:57"
        s = serial.strip()
        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y/%m/%d %H:%M:%S"]:
            try:
                return datetime.strptime(s[:19] if len(s)>=19 else s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        # If it starts with YYYY-MM-DD, just take the first 10 chars
        if len(s) >= 10 and s[4] == '-' and s[7] == '-':
            return s[:10]
    if isinstance(serial, (int, float)) and serial > 40000:
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(serial))).strftime("%Y-%m-%d")
    return None

def in_range(date_str):
    """Check if date string is within range. date_str is YYYY-MM-DD format."""
    if not date_str:
        return False
    return START_DATE <= date_str <= END_DATE

def date_to_month(date_str):
    return date_str[:7]

print("Loading workbook...", file=sys.stderr)
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# ============================================================
# 0. Dynamic brand discovery + END_DATE auto-detection
# ============================================================
ws_orders_brand = wb["交易订单"]
brand_col_ord = 8
raw_brands = set()
all_dates_set = set()
for r in range(2, min(ws_orders_brand.max_row + 1, 5000)):
    b = ws_orders_brand.cell(r, brand_col_ord).value
    if b is not None: raw_brands.add(str(b).strip())
    d = excel_to_date(ws_orders_brand.cell(r, 21).value) # 订单状态 col
    if d: all_dates_set.add(d)
END_DATE = max(all_dates_set) if all_dates_set else "2026-07-28"

DYNAMIC_BRANDS = set()
for b in raw_brands:
    DYNAMIC_BRANDS.add(brand_short(standardize_brand(b)))
ALL_BRANDS_LIST = sorted(DYNAMIC_BRANDS)
print(f"Discovered {len(raw_brands)} raw brands → {len(ALL_BRANDS_LIST)} standard: {ALL_BRANDS_LIST}", file=sys.stderr)
print(f"Date range: {START_DATE} ~ {END_DATE}", file=sys.stderr)
# ============================================================

# ============================================================
# 1. 大盘指数 (日韩表 + 单肩包)
# ============================================================
print("1. Extracting 大盘指数...", file=sys.stderr)
ws = wb["大盘指数"]
market_data = []          # 日韩表大盘
market_monthly = defaultdict(list)
market_data_bag = []      # 单肩包大盘
market_monthly_bag = defaultdict(list)
for r in range(2, ws.max_row + 1):
    d = excel_to_date(ws.cell(r, 2).value)
    v = ws.cell(r, 3).value  # 日韩表大盘成交GMV指数
    v_bag = ws.cell(r, 4).value  # 单肩包大盘成交GMV指数
    if d and in_range(d):
        if v:
            market_data.append({"date": d, "value": round(float(v), 2)})
            market_monthly[date_to_month(d)].append(float(v))
        if v_bag:
            market_data_bag.append({"date": d, "value": round(float(v_bag), 2)})
            market_monthly_bag[date_to_month(d)].append(float(v_bag))

market_monthly_avg = {m: round(sum(vals)/len(vals), 2) for m, vals in sorted(market_monthly.items())}
market_monthly_bag_avg = {m: round(sum(vals)/len(vals), 2) for m, vals in sorted(market_monthly_bag.items())}
print(f"  Market(日韩表): {len(market_data)} daily records, {len(market_monthly_avg)} months", file=sys.stderr)
print(f"  Market(单肩包): {len(market_data_bag)} daily records, {len(market_monthly_bag_avg)} months", file=sys.stderr)
for m in sorted(set(list(market_monthly_avg.keys()) + list(market_monthly_bag_avg.keys()))):
    a1 = market_monthly_avg.get(m, "—")
    a2 = market_monthly_bag_avg.get(m, "—")
    print(f"    {m}: 日韩表={a1}, 单肩包={a2}", file=sys.stderr)

# ============================================================
# 2. 货盘表 - build SPU -> brand mapping
# ============================================================
print("2. Building SPU->brand mapping...", file=sys.stderr)
ws = wb["货盘表"]
spu_brand = {}
spu_name = {}
for r in range(2, ws.max_row + 1):
    spuid = ws.cell(r, 1).value
    name = ws.cell(r, 2).value
    brand = ws.cell(r, 5).value
    if spuid and brand is not None:
        std_brand = standardize_brand(brand)
        try:
            spu_brand[int(spuid)] = std_brand
            spu_name[int(spuid)] = str(name)
        except (ValueError, TypeError):
            pass

print(f"  SPU mapping: {len(spu_brand)} SPUs", file=sys.stderr)

# ============================================================
# 3. 交易订单 - 按月/按SPU汇总
# ============================================================
print("3. Extracting 交易订单...", file=sys.stderr)
ws = wb["交易订单"]
# Headers: 订单号, 订单类型, spuID, skuID, 商品名称, 货号, SKU货号, 品牌, 规格, 数量, 出价金额(元), ...
# Find column indices
header = {}
for c in range(1, ws.max_column + 1):
    h = ws.cell(1, c).value
    if h: header[str(h).strip()] = c

col_order_status = None
col_pay_time = None
col_order_time = None
for c in range(1, ws.max_column + 1):
    h = str(ws.cell(1, c).value or "").strip()
    if "订单状态" in h:
        col_order_status = c
    elif "支付时间" in h or "付款时间" in h:
        col_pay_time = c
    elif "订单创建时间" in h or "下单时间" in h:
        col_order_time = c

print(f"  order_status col={col_order_status}, pay_time={col_pay_time}, order_time={col_order_time}", file=sys.stderr)

spu_id_col = 3  # spuID
sku_id_col = 4  
name_col = 5
brand_col = 8
amount_col = 11
qty_col = 10
huohao_col = 6  # 货号

# ============================================================
# v2.1: 统一合约入口 — 状态集从 store_config 读取，禁止硬编码
# ============================================================
import sys as _sys
_sys.path.insert(0, '/home/Vic/.hermes/skills/dewu/dewu-operations-analysis/scripts')
import contract_lib as _cl
_CFG = _cl.load_store_config('喜过')
VALID_STATUSES = list(_cl.get_paid_states(_CFG))
EXCLUDED_STATUSES = list(_cl.get_excluded_states(_CFG))
# ============================================================

orders_monthly = defaultdict(lambda: defaultdict(lambda: {"gmv": 0.0, "orders": 0}))
orders_by_spu = defaultdict(lambda: defaultdict(lambda: {"gmv": 0.0, "orders": 0}))
orders_by_huohao = defaultdict(lambda: defaultdict(lambda: {"gmv": 0.0, "orders": 0}))  # 按货号汇总

processed = 0
for r in range(2, ws.max_row + 1):
    # C5: 排除非订单汇总行（"汇总""合计"等Excel末行）
    oid_val = ws.cell(r, 1).value
    if oid_val and any(kw in str(oid_val) for kw in ['汇总', '合计', '小计', '总计']):
        continue

    spuid = ws.cell(r, spu_id_col).value
    amount = ws.cell(r, amount_col).value
    brand = ws.cell(r, brand_col).value
    status = ws.cell(r, col_order_status).value if col_order_status else None
    
    # Get date
    date_str = None
    if col_pay_time:
        dt = ws.cell(r, col_pay_time).value
        date_str = excel_to_date(dt)
    if not date_str and col_order_time:
        dt = ws.cell(r, col_order_time).value
        date_str = excel_to_date(dt)
    
    if not (spuid and amount and status in VALID_STATUSES and date_str and in_range(date_str)):
        continue
    std_b = standardize_brand(brand)
    b = brand_short(std_b)
    month = date_to_month(date_str)
    gmv = float(amount) * (ws.cell(r, qty_col).value or 1)
    
    orders_monthly[b][month]["gmv"] += gmv
    orders_monthly[b][month]["orders"] += 1
    
    spu_key = spu_name.get(int(spuid), str(spuid))
    orders_by_spu[b][spu_key]["gmv"] += gmv
    orders_by_spu[b][spu_key]["orders"] += 1
    orders_by_spu[b][spu_key]["spuid"] = int(spuid)
    
    # 按货号汇总
    hh = str(ws.cell(r, huohao_col).value).strip() if ws.cell(r, huohao_col).value else spu_key
    orders_by_huohao[b][hh]["gmv"] += gmv
    orders_by_huohao[b][hh]["orders"] += 1
    
    processed += 1

print(f"  Processed {processed} valid orders", file=sys.stderr)

# Convert to sorted lists
orders_monthly_out = {}
for brand in ALL_BRANDS_LIST:
    orders_monthly_out[brand] = {}
    for m in sorted(orders_monthly[brand].keys()):
        orders_monthly_out[brand][m] = {
            "gmv": round(orders_monthly[brand][m]["gmv"], 2),
            "orders": orders_monthly[brand][m]["orders"]
        }

orders_by_spu_out = {}
for brand in ALL_BRANDS_LIST:
    spus = []
    for spu_name_key, data in sorted(orders_by_spu[brand].items(), key=lambda x: -x[1]["gmv"]):
        spus.append({
            "spu": spu_name_key,
            "spuid": data["spuid"],
            "gmv": round(data["gmv"], 2),
            "orders": data["orders"]
        })
    orders_by_spu_out[brand] = spus

# ============================================================
# 4. 商详访客数据 - UV按月/按SPU
# ============================================================
print("4. Extracting 商详访客数据...", file=sys.stderr)
ws = wb["商详访客数据"]
# Columns: 月份, 日期, SPUID, 商品名称, 货号, 品牌名称, 类目名称, 商品类型, 出价渠道, 支付订单金额, 支付订单量, ...
# Find the key columns
uv_header = {}
for c in range(1, min(50, ws.max_column + 1)):
    h = ws.cell(1, c).value
    if h: uv_header[str(h).strip()] = c

# Check what columns exist
print(f"  商详 headers (first 20 cols): {[ws.cell(1,c).value for c in range(1,21)]}", file=sys.stderr)

# Find indices: SPUID=col3, 品牌=col6, 商详访客人数 column
spuid_col_uv = 3
brand_col_uv = 6
# Look for 商详访客人数 column
uv_col = None
pay_amount_col = None
pay_orders_col = None
for c in range(1, min(50, ws.max_column + 1)):
    h = str(ws.cell(1, c).value or "").strip()
    # Specifically look for 商详访问指数（UV）or similar
    if "商详访问" in h and ("uv" in h.lower() or "指数" in h):
        if uv_col is None:
            uv_col = c
    elif "支付订单金额" in h:
        pay_amount_col = c
    elif "支付订单量" in h:
        pay_orders_col = c

print(f"  UV col={uv_col}, pay_amount={pay_amount_col}, pay_orders={pay_orders_col}", file=sys.stderr)

uv_monthly = defaultdict(lambda: defaultdict(lambda: {"uv": 0, "gmv": 0.0, "orders": 0}))
uv_by_spu = defaultdict(lambda: defaultdict(lambda: {"daily": [], "monthly": defaultdict(lambda: {"uv": 0, "gmv": 0.0, "orders": 0})}))

uv_processed = 0
for r in range(2, ws.max_row + 1):
    spuid = ws.cell(r, spuid_col_uv).value
    brand = ws.cell(r, brand_col_uv).value
    
    # Get date - column 2 is 日期
    date_cell = ws.cell(r, 2).value
    date_str = excel_to_date(date_cell)
    
    if not (spuid and brand is not None and date_str and in_range(date_str)):
        continue
    
    b = brand_short(standardize_brand(brand))
    month = date_to_month(date_str)
    
    uv = ws.cell(r, uv_col).value if uv_col else 0
    pay_amt = ws.cell(r, pay_amount_col).value if pay_amount_col else 0
    pay_ord = ws.cell(r, pay_orders_col).value if pay_orders_col else 0
    
    def clean_num(v):
        """Extract number from values like '122 (指数)' or '0 (指数)'"""
        if v is None:
            return 0
        if isinstance(v, (int, float)):
            return v
        if isinstance(v, str):
            # Take only the part before space/parenthesis
            import re
            m = re.match(r'[\d.]+', v.strip())
            return float(m.group()) if m else 0
        return 0
    
    uv_val = int(clean_num(uv))
    pay_amt_val = float(clean_num(pay_amt))
    pay_ord_val = int(clean_num(pay_ord))
    
    uv_monthly[b][month]["uv"] += uv_val
    uv_monthly[b][month]["gmv"] += pay_amt_val
    uv_monthly[b][month]["orders"] += pay_ord_val
    
    spu_key = spu_name.get(int(spuid), str(spuid))
    uv_by_spu[b][spu_key]["daily"].append({
        "date": date_str,
        "uv": uv_val,
        "gmv": round(pay_amt_val, 2),
        "orders": pay_ord_val
    })
    uv_by_spu[b][spu_key]["monthly"][month]["uv"] += uv_val
    uv_by_spu[b][spu_key]["monthly"][month]["gmv"] += pay_amt_val
    uv_by_spu[b][spu_key]["monthly"][month]["orders"] += pay_ord_val
    
    uv_processed += 1

print(f"  UV processed {uv_processed} records", file=sys.stderr)

uv_monthly_out = {}
for brand in ALL_BRANDS_LIST:
    uv_monthly_out[brand] = {}
    for m in sorted(uv_monthly[brand].keys()):
        uv_monthly_out[brand][m] = {
            "uv": uv_monthly[brand][m]["uv"],
            "gmv": round(uv_monthly[brand][m]["gmv"], 2),
            "orders": uv_monthly[brand][m]["orders"]
        }

uv_by_spu_out = {}
for brand in ALL_BRANDS_LIST:
    uv_by_spu_out[brand] = {}
    for spu_key, data in uv_by_spu[brand].items():
        uv_by_spu_out[brand][spu_key] = {
            "daily": sorted(data["daily"], key=lambda x: x["date"]),
            "monthly": {m: {"uv": d["uv"], "gmv": round(d["gmv"], 2), "orders": d["orders"]} 
                       for m, d in sorted(data["monthly"].items())}
        }

# ============================================================
# 5. 得物推数据 - 按月汇总（双源品牌匹配：货盘表 + 交易订单）
# ============================================================
print("5. Extracting 得物推数据...", file=sys.stderr)
ws = wb["得物推数据-商品"]
# Headers (0-indexed): 时间, 用户ID, 计划名称, 计划ID, 计划类型, 优化目标, 商品ID2, 消耗(元), ...
# openpyxl 1-indexed: col 1=时间, col 7=商品ID2, col 8=消耗(元), col 15=直接支付单量, col 16=直接支付金额, col 17=引导支付单量, col 18=引导支付金额
push_header = {}
for c in range(1, min(30, ws.max_column + 1)):
    h = ws.cell(1, c).value
    if h: push_header[str(h).strip()] = c

time_col = 1
goods_id_col = 7
cost_col = 8
direct_orders_col = 15
direct_gmv_col = 16
indirect_orders_col = 17
indirect_gmv_col = 18
# Note: this sheet has NO brand/huohao columns — use SPU mapping from 货盘表
print(f"  Push: time_col={time_col}, goods_id_col={goods_id_col}, cost_col={cost_col}", file=sys.stderr)

# Build SPU->brand mapping from 交易订单 too (supplement 货盘表)
spu_brand_from_orders = {}
ws_orders = wb["交易订单"]
spu_id_col_ord = 3
brand_col_ord = 8
for r in range(2, ws_orders.max_row + 1):
    spuid = ws_orders.cell(r, spu_id_col_ord).value
    brand = ws_orders.cell(r, brand_col_ord).value
    if spuid and brand is not None:
        try:
            spu_brand_from_orders[int(spuid)] = brand_short(standardize_brand(brand))
        except (ValueError, TypeError):
            pass
print(f"  Orders SPU→brand mapping: {len(spu_brand_from_orders)} SPUs", file=sys.stderr)

push_monthly = defaultdict(lambda: defaultdict(float))
push_daily = []  # Per-day per-goods push records for Module 5
push_processed = 0
for r in range(2, ws.max_row + 1):
    dt = ws.cell(r, time_col).value
    date_str = excel_to_date(dt)
    good_id = ws.cell(r, goods_id_col).value
    cost = ws.cell(r, cost_col).value
    
    if not (date_str and good_id and cost and in_range(date_str)):
        continue
    
    # Match brand: use SPU mapping from 货盘表 / 交易订单
    brand = None
    huohao = None
    try:
        gid = int(good_id)
        brand = spu_brand.get(gid) or spu_brand_from_orders.get(gid)
        huohao = spu_name.get(gid)
    except (ValueError, TypeError):
        pass
    
    if not brand:
        continue
    
    if not huohao:
        huohao = str(good_id)
    
    month = date_to_month(date_str)
    cost_val = float(cost) if cost else 0
    
    push_monthly[brand][month] += cost_val
    
    # Build daily record for Module 5
    def safe_float(v):
        try: return float(v) if v else 0.0
        except: return 0.0
    def safe_int(v):
        try: return int(float(v)) if v else 0
        except: return 0
    
    push_daily.append({
        "date": date_str,
        "brand": brand,
        "huohao": huohao,
        "cost": round(cost_val, 2),
        "direct_orders": safe_int(ws.cell(r, direct_orders_col).value),
        "direct_gmv": round(safe_float(ws.cell(r, direct_gmv_col).value), 2),
        "indirect_orders": safe_int(ws.cell(r, indirect_orders_col).value),
        "indirect_gmv": round(safe_float(ws.cell(r, indirect_gmv_col).value), 2)
    })
    push_processed += 1

print(f"  Push processed {push_processed} records (daily: {len(push_daily)})", file=sys.stderr)

push_monthly_out = {}
for brand in ALL_BRANDS_LIST:
    push_monthly_out[brand] = {m: round(v, 2) for m, v in sorted(push_monthly[brand].items())}

# ============================================================
# 6. 社区投放任务 - 按月汇总
# ============================================================
print("6. Extracting 社区投放任务...", file=sys.stderr)
ws = wb["社区投放任务"]
# Headers: 任务月份, 父任务ID, 子任务ID, 任务名称, 任务发布时间, 任务推广形式, 任务模式, 任务状态,
#   任务完成时间, 任务金额, 实际任务金额, 合作达人, 达人uid, ... 匹配货号(29)
comm_header = {}
for c in range(1, min(35, ws.max_column + 1)):
    h = ws.cell(1, c).value
    if h: comm_header[str(h).strip()] = c

# Find columns
task_month_col = comm_header.get("任务月份", 1)
match_goods_col = comm_header.get("匹配货号", 29)
amount_col_comm = comm_header.get("实际任务金额", 11)
time_col_comm = comm_header.get("动态发布时间", 19)
status_col = comm_header.get("任务状态", 8)

print(f"  Comm cols: month={task_month_col}, match_goods={match_goods_col}, amount={amount_col_comm}, time={time_col_comm}, status={status_col}", file=sys.stderr)

comm_monthly = defaultdict(lambda: defaultdict(lambda: {"cost": 0.0, "tasks": 0}))
comm_processed = 0
for r in range(2, ws.max_row + 1):
    match_goods = ws.cell(r, match_goods_col).value
    amount_str = ws.cell(r, amount_col_comm).value
    dt = ws.cell(r, time_col_comm).value
    date_str = excel_to_date(dt)
    
    if not (match_goods and amount_str and date_str and in_range(date_str)):
        continue
    
    # Parse amount
    try:
        if isinstance(amount_str, str):
            amount_str = amount_str.replace("¥", "").replace("元", "").replace(",", "").strip()
        amount = float(amount_str)
    except:
        continue
    
    # Match brand from goods name keywords
    goods_list = str(match_goods).split(",")
    brands_for_task = set()
    for g in goods_list:
        g = g.strip()
        # CASIO keywords
        if any(kw in g for kw in ['CASIO', '卡西欧', 'GA-', 'W-', 'F-', 'A1', 'AE-', 'WS-', 'DW-', 'LTP']):
            brands_for_task.add('卡西欧')
        # COACH keywords
        elif any(kw in g for kw in ['COACH', '蔻驰', 'Coach']):
            brands_for_task.add('蔻驰')
    
    if not brands_for_task:
        continue
    
    # Split cost evenly among brands
    month = date_to_month(date_str)
    per_brand = amount / len(brands_for_task)
    for b in brands_for_task:
        comm_monthly[b][month]["cost"] += per_brand
        comm_monthly[b][month]["tasks"] += 1
    
    comm_processed += 1

print(f"  Comm processed {comm_processed} records", file=sys.stderr)

comm_monthly_out = {}
for brand in ALL_BRANDS_LIST:
    comm_monthly_out[brand] = {}
    for m in sorted(comm_monthly[brand].keys()):
        comm_monthly_out[brand][m] = {
            "cost": round(comm_monthly[brand][m]["cost"], 2),
            "tasks": comm_monthly[brand][m]["tasks"]
        }

# ============================================================
# 6b. 社区投放任务 - 按货号明细（新增模块七数据）
# ============================================================
print("6b. Extracting 社区投放任务-货号明细...", file=sys.stderr)
comm_tasks = []

for r in range(2, ws.max_row + 1):
    # 任务月份 (col 1) - may be Excel serial or datetime
    task_month_val = ws.cell(r, 1).value
    task_month = excel_to_date(task_month_val)
    if task_month:
        task_month = task_month[:7]  # YYYY-MM
    
    # 动态发布时间 (col 19)
    pub_date_val = ws.cell(r, 19).value
    pub_date = excel_to_date(pub_date_val)
    
    # 实际任务金额 (col 11)
    amount_val = ws.cell(r, 11).value
    if amount_val is None:
        continue
    try:
        if isinstance(amount_val, str):
            amount_val = amount_val.replace("¥", "").replace("元", "").replace(",", "").strip()
        amount = float(amount_val)
    except:
        continue
    
    # 曝光 (col 21), 阅读数 (col 22), 商详访问 (col 24)
    def safe_int_comm(v):
        if v is None:
            return 0
        if isinstance(v, (int, float)):
            return int(v)
        s = str(v).strip().replace(",", "")
        if s == "暂无" or s == "":
            return 0
        try:
            return int(float(s))
        except:
            return 0
    
    exposure = safe_int_comm(ws.cell(r, 21).value)
    reads = safe_int_comm(ws.cell(r, 22).value)
    visits = safe_int_comm(ws.cell(r, 24).value)
    
    # 匹配货号1 (col 29) and 匹配货号2 (col 30)
    match_goods_1 = ws.cell(r, 29).value
    match_goods_2 = ws.cell(r, 30).value
    
    all_goods = []
    if match_goods_1:
        for g in str(match_goods_1).split(","):
            g = g.strip()
            if g and g != "暂无" and g != "无":
                all_goods.append(g)
    if match_goods_2:
        for g in str(match_goods_2).split(","):
            g = g.strip()
            if g and g != "暂无" and g != "无":
                all_goods.append(g)
    
    if not all_goods:
        continue
    
    # Determine brand per goods item
    for goods in all_goods:
        brand = None
        if any(kw in goods for kw in ['CASIO', '卡西欧', 'GA-', 'W-', 'F-', 'A1', 'AE-', 'WS-', 'DW-', 'LTP']):
            brand = '卡西欧'
        elif any(kw in goods for kw in ['COACH', '蔻驰', 'Coach']):
            brand = '蔻驰'
        
        if not brand:
            continue
        
        # Split amount evenly among goods in this task
        per_amount = round(amount / len(all_goods), 2)
        
        comm_tasks.append({
            "brand": brand,
            "huohao": goods,
            "month": task_month or "",
            "pub_date": pub_date or "",
            "amount": per_amount,
            "exposure": exposure,
            "reads": reads,
            "visits": visits
        })

print(f"  Comm task records (by huohao): {len(comm_tasks)}", file=sys.stderr)
# Show sample
sample_brands = {}
for t in comm_tasks:
    b = t["brand"]
    if b not in sample_brands:
        sample_brands[b] = 0
    sample_brands[b] += 1
for b, c in sample_brands.items():
    print(f"    {b}: {c} records", file=sys.stderr)

# ============================================================
# 7. Build output (supplemental fields added separately)
# ============================================================
months_list = sorted(set(
    list(orders_monthly_out.get("卡西欧", {}).keys()) +
    list(orders_monthly_out.get("蔻驰", {}).keys()) +
    list(market_monthly_avg.keys())
))

output = {
    "dateRange": {"start": START_DATE, "end": END_DATE},
    "months": months_list,
    "orders_monthly": orders_monthly_out,
    "orders_by_spu": orders_by_spu_out,
    "uv_monthly": uv_monthly_out,
    "uv_by_spu": uv_by_spu_out,
    "market": market_data,
    "market_monthly_avg": market_monthly_avg,
    "market_bag": market_data_bag,
    "market_monthly_bag_avg": market_monthly_bag_avg,
    "push_monthly": push_monthly_out,
    "push_daily": push_daily,
    "comm_monthly": comm_monthly_out,
    "comm_tasks": comm_tasks
}

# Write as data.js (SAFE ATOMIC VERSION — 2026-07-30 authorized)
OUTPUT_PATH = os.path.join(_OUTPUT_DIR, "data.js") if _OUTPUT_DIR else "/home/Vic/dewu-reports/XiguoAnalysis/2026/data.js"

# Step 1: 序列化到临时文件并落盘
import hashlib as _hashlib
json_str = json.dumps(output, ensure_ascii=False, separators=(",", ":"))
_file_content = "var D=" + json_str + ";\n"
tmp_path = OUTPUT_PATH + ".tmp"

with open(tmp_path, "w", encoding="utf-8") as f:
    f.write(_file_content)
    f.flush()
    __import__('os').fsync(f.fileno())

# Step 2: 重新读取并验证完整性
with open(tmp_path, "r", encoding="utf-8") as f:
    _written = f.read()

_bo = _written.count("{")
_bc = _written.count("}")
assert _bo == _bc, f"FATAL: data.js bracket mismatch {{{_bo} vs {_bc}}}"
assert _written.startswith("var D="), "FATAL: data.js missing var D= prefix"
assert _written.rstrip().endswith("};"), f"FATAL: data.js bad terminator: {repr(_written[-20:])}"

# Step 3: 验证JSON载荷
_json_extracted = _written[_written.index("{"):_written.rindex("};")+1]
_parsed = json.loads(_json_extracted)
assert "months" in _parsed, "FATAL: data.js missing months field"
assert "orders_monthly" in _parsed, "FATAL: data.js missing orders_monthly field"

# Step 4: 原子替换
_sha = _hashlib.sha256(_written.encode("utf-8")).hexdigest()
__import__('os').rename(tmp_path, OUTPUT_PATH)

print(f"\n✅ Written to {OUTPUT_PATH}", file=sys.stderr)
print(f"  Size: {len(json_str)} chars, File: {len(_written)} bytes", file=sys.stderr)
print(f"  SHA256: {_sha[:16]}...", file=sys.stderr)
print(f"  Brackets: {_bo}/{_bc} OK, JSON valid: OK", file=sys.stderr)
print(f"  Market records: {len(market_data)}", file=sys.stderr)
print(f"  Months: {months_list}", file=sys.stderr)

# ⚠️ IMPORTANT: data.js generated by openpyxl above is MISSING three fields
# needed by app.js 模块二/三/四/五:
#   - uv_daily_by_brand
#   - uv_spu_data
#   - huohao_daily
# Run the pandas supplement script to inject these:
#   python3 /home/Vic/dewu-reports/XiguoAnalysis/supplement_data.py

# ============================================================
# 8. 五分类提取 (extract_fivecat)
# ============================================================
def extract_fivecat(store_name='喜过'):
    """
    从交易订单和售后订单中提取五分类数据：
    正常、退货、取消、不明确 四个分类的GMV/订单数
    以及成熟退货率等KPI指标
    """
    import sys as _sys
    _sys.path.insert(0, '/home/Vic/dewu-reports')
    from extract_fivecat import classify_store, STORE_CONFIG
    fc = classify_store(store_name, STORE_CONFIG[store_name])
    
    fivecat = {
        'n_pay': fc['n_pay'],
        'gmv_pay': fc['gmv_pay'],
        'mature_return_rate': fc['mature_return_rate'],
        'as_start': fc['as_start'],
        'n_ret': fc['n_ret'],
        'gmv_ret': fc['gmv_ret'],
        'n_cancel': fc['n_cancel'],
        'gmv_cancel': fc['gmv_cancel'],
        'n_unclear': fc['n_unclear'],
        'gmv_unclear': fc['gmv_unclear'],
        'n_normal': fc['n_normal'],
        'gmv_normal': fc['gmv_normal'],
        'refund_total': fc['refund_total'],
        'n_reliable_mature': fc['n_reliable_mature'],
        'n_reliable_ret': fc['n_reliable_ret'],
        'daily_gmv': fc['daily_gmv'],
    }
    
    print(f"  五分类: pay={fivecat['n_pay']}, gmv={fivecat['gmv_pay']:.0f}, "
          f"mature_rate={fivecat['mature_return_rate']}%, "
          f"ret={fivecat['n_ret']}, cancel={fivecat['n_cancel']}, "
          f"unclear={fivecat['n_unclear']}, normal={fivecat['n_normal']}", file=_sys.stderr)
    
    return fivecat

# Extract and write fivedata.js
print("\n8. Extracting 五分类...", file=sys.stderr)
FIVECAT = extract_fivecat('喜过')

FIVECAT_PATH = os.path.join(_OUTPUT_DIR, "fivedata.js") if _OUTPUT_DIR else "/home/Vic/dewu-reports/XiguoAnalysis/2026/fivedata.js"
with open(FIVECAT_PATH, "w", encoding="utf-8") as f:
    f.write("const FIVECAT = ")
    json.dump(FIVECAT, f, ensure_ascii=False, separators=(",", ":"))
    f.write(";\n")

print(f"✅ Written fivedata.js to {FIVECAT_PATH}", file=sys.stderr)
print(f"  Size: {len(json.dumps(FIVECAT, ensure_ascii=False))} chars", file=sys.stderr)
