#!/usr/bin/env python3
"""
C3 — 喜过包包大盘数据生成
============================
从飞书 喜过数据源收集表.xlsx 的「大盘指数」Sheet 中提取
「单肩包大盘成交GMV指数」日粒度数据，生成 page-consumable
market_bag_data.js 文件。

输出:
  MKT_BAG          — [{date, value}, ...]  日粒度
  MKT_BAG_MONTHLY  — {month: avg_value, ...}  月粒度（对象格式，匹配页面消费者）

用法:
  python3 generate_market_bag.py --source-file <source.xlsx> [--output-dir DIR]
  python3 generate_market_bag.py <source.xlsx> [--output-dir DIR]
  
  默认输出到当前目录，--output-dir 指定隔离目录。
  使用临时文件 + rename 原子替换。
  不影响 OID/GMV/品牌/SPUID 数据。
"""
import json
import os
import sys
import argparse
import hashlib
from collections import defaultdict
from datetime import datetime

def date_to_month(d):
    """'2026-07-30' → '2026-07'"""
    if isinstance(d, datetime):
        return d.strftime('%Y-%m')
    return str(d)[:7]

def main():
    parser = argparse.ArgumentParser(description='Generate 喜过 包包大盘数据')
    parser.add_argument('source', nargs='?', default=None, help='Path to 喜过数据源收集表.xlsx (positional)')
    parser.add_argument('--source-file', default=None, help='Path to 喜过数据源收集表.xlsx (named)')
    parser.add_argument('--output-dir', default=None, help='Output directory (default: cwd)')
    args = parser.parse_args()
    
    # Resolve source: --source-file takes priority over positional
    source_path = args.source_file or args.source
    if not source_path:
        print("ERROR: --source-file or positional source required", file=sys.stderr)
        sys.exit(1)
    if not os.path.exists(source_path):
        print(f"ERROR: source file not found: {source_path}", file=sys.stderr)
        sys.exit(1)
    
    out_dir = args.output_dir or '.'
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, 'market_bag_data.js')
    
    # ── Read source ──
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required", file=sys.stderr)
        sys.exit(1)
    
    wb = openpyxl.load_workbook(source_path, data_only=True, read_only=True)
    
    if '大盘指数' not in wb.sheetnames:
        print(f"ERROR: '大盘指数' sheet not found in {source_path}", file=sys.stderr)
        print(f"Available sheets: {wb.sheetnames}", file=sys.stderr)
        wb.close()
        sys.exit(1)
    
    ws = wb['大盘指数']
    
    # ── Extract 单肩包 daily data ──
    market_bag = []
    monthly_vals = defaultdict(list)
    
    row_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        if row[1] is None:
            break
        date_val = row[1]  # 日期 column
        bag_val = row[3]   # 单肩包大盘成交GMV指数 column
        
        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            date_str = str(date_val)[:10]
        
        if bag_val is not None:
            try:
                v = round(float(bag_val), 2)
                market_bag.append({"date": date_str, "value": v})
                monthly_vals[date_to_month(date_str)].append(v)
                row_count += 1
            except (ValueError, TypeError):
                pass
    
    wb.close()
    
    if row_count == 0:
        print("ERROR: No valid 单肩包 data extracted", file=sys.stderr)
        sys.exit(1)
    
    print(f"Extracted {row_count} daily 单肩包 records", file=sys.stderr)
    print(f"  Range: {market_bag[0]['date']} ~ {market_bag[-1]['date']}", file=sys.stderr)
    
    # ── Monthly aggregation (as object, matching page consumer format) ──
    market_monthly_bag = {}
    for m in sorted(monthly_vals.keys()):
        vals = monthly_vals[m]
        market_monthly_bag[m] = round(sum(vals) / len(vals), 2)
    
    print(f"  Monthly: {len(market_monthly_bag)} months", file=sys.stderr)
    print(f"  Range: {list(market_monthly_bag.keys())[0]} ~ {list(market_monthly_bag.keys())[-1]}", file=sys.stderr)
    
    # ── Build output ──
    bag_js = json.dumps(market_bag, ensure_ascii=False, separators=(',', ':'))
    monthly_js = json.dumps(market_monthly_bag, ensure_ascii=False, separators=(',', ':'))
    
    content = (
        "// 喜过 包包大盘数据 (单肩包大盘成交GMV指数)\n"
        f"// Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"// Source: 喜过数据源收集表.xlsx → 大盘指数 Sheet\n"
        f"const MKT_BAG = {bag_js};\n"
        f"const MKT_BAG_MONTHLY = {monthly_js};\n"
    )
    
    # ── Atomic write via tmp file ──
    tmp_path = out_file + '.tmp'
    with open(tmp_path, 'w', encoding='utf-8') as f:
        f.write(content)
        f.flush()
        os.fsync(f.fileno())
    
    # Verify
    with open(tmp_path, 'r', encoding='utf-8') as f:
        written = f.read()
    
    assert 'const MKT_BAG = ' in written, "FATAL: MKT_BAG missing"
    assert 'const MKT_BAG_MONTHLY = ' in written, "FATAL: MKT_BAG_MONTHLY missing"
    
    sha = hashlib.sha256(written.encode('utf-8')).hexdigest()
    
    os.rename(tmp_path, out_file)
    
    size = os.path.getsize(out_file)
    print(f"\n✅ Written {out_file}", file=sys.stderr)
    print(f"   Size: {size} bytes, SHA256: {sha}", file=sys.stderr)
    print(f"   MKT_BAG: {len(market_bag)} daily records", file=sys.stderr)
    print(f"   MKT_BAG_MONTHLY: {len(market_monthly_bag)} monthly records", file=sys.stderr)
    
    return 0

if __name__ == '__main__':
    sys.exit(main())
