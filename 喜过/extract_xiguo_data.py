"""
喜过分析模板 - 数据提取脚本
从 喜过数据源收集表.xlsx 提取卡西欧 vs 蔻驰的全量运营数据
输出: data.js (内联到看板HTML)

核心修复:
1. 大盘指数提取覆盖全时间范围（不只是May）
2. 社区投放/得物推按月汇总
3. UV月度数据
4. SPU级别的UV、支付订单、支付金额详情
"""
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict
import json
import sys

EXCEL_PATH = "/mnt/e/Obsidian本地仓库/09-数据源/喜过数据源收集表.xlsx"
START_DATE = "2026-01-01"
END_DATE = "2026-05-14"
BRANDS = ["CASIO/卡西欧", "COACH/蔻驰"]
BRAND_SHORT = {"CASIO/卡西欧": "卡西欧", "COACH/蔻驰": "蔻驰"}

def excel_to_date(serial):
    """Convert Excel date serial or string to YYYY-MM-DD string"""
    if isinstance(serial, datetime):
        return serial.strftime("%Y-%m-%d")
    if isinstance(serial, str) and serial.strip():
        # Try to parse string dates like "2026-05-07 14:21:57"
        s = serial.strip()
        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y/%m/%d %H:%M:%S"]:
            try:
                return datetime.strptime(s[:19] if len(s)>=19 else s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        # If it starts with YYYY-MM-DD, just take the first 10 chars
        if len(s) >= 10 and s[4] == '-' and s[7] == '-':
            return s[:10]
    if isinstance(serial, (int, float)) and serial > 40000:
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=int(serial))).strftime("%Y-%m-%d")
    return None

def in_range(date_str):
    """Check if date string is within range. date_str is YYYY-MM-DD format."""
    if not date_str:
        return False
    return START_DATE <= date_str <= END_DATE

def date_to_month(date_str):
    return date_str[:7]

print("Loading workbook...", file=sys.stderr)
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# ============================================================
# 1. 大盘指数 (日韩表 + 单肩包)
# ============================================================
print("1. Extracting 大盘指数...", file=sys.stderr)
ws = wb["大盘指数"]
market_data = []          # 日韩表大盘
market_monthly = defaultdict(list)
market_data_bag = []      # 单肩包大盘
market_monthly_bag = defaultdict(list)
for r in range(2, ws.max_row + 1):
    d = excel_to_date(ws.cell(r, 2).value)
    v = ws.cell(r, 3).value  # 日韩表大盘成交GMV指数
    v_bag = ws.cell(r, 4).value  # 单肩包大盘成交GMV指数
    if d and in_range(d):
        if v:
            market_data.append({"date": d, "value": round(float(v), 2)})
            market_monthly[date_to_month(d)].append(float(v))
        if v_bag:
            market_data_bag.append({"date": d, "value": round(float(v_bag), 2)})
            market_monthly_bag[date_to_month(d)].append(float(v_bag))

market_monthly_avg = {m: round(sum(vals)/len(vals), 2) for m, vals in sorted(market_monthly.items())}
market_monthly_bag_avg = {m: round(sum(vals)/len(vals), 2) for m, vals in sorted(market_monthly_bag.items())}
print(f"  Market(日韩表): {len(market_data)} daily records, {len(market_monthly_avg)} months", file=sys.stderr)
print(f"  Market(单肩包): {len(market_data_bag)} daily records, {len(market_monthly_bag_avg)} months", file=sys.stderr)
for m in sorted(set(list(market_monthly_avg.keys()) + list(market_monthly_bag_avg.keys()))):
    a1 = market_monthly_avg.get(m, "—")
    a2 = market_monthly_bag_avg.get(m, "—")
    print(f"    {m}: 日韩表={a1}, 单肩包={a2}", file=sys.stderr)

# ============================================================
# 2. 货盘表 - build SPU -> brand mapping
# ============================================================
print("2. Building SPU->brand mapping...", file=sys.stderr)
ws = wb["货盘表"]
spu_brand = {}
spu_name = {}
for r in range(2, ws.max_row + 1):
    spuid = ws.cell(r, 1).value
    name = ws.cell(r, 2).value
    brand = ws.cell(r, 5).value
    if spuid and brand in BRANDS:
        try:
            spu_brand[int(spuid)] = BRAND_SHORT[brand]
            spu_name[int(spuid)] = str(name)
        except (ValueError, TypeError):
            pass

print(f"  SPU mapping: {len(spu_brand)} SPUs", file=sys.stderr)

# ============================================================
# 3. 交易订单 - 按月/按SPU汇总
# ============================================================
print("3. Extracting 交易订单...", file=sys.stderr)
ws = wb["交易订单"]
# Headers: 订单号, 订单类型, spuID, skuID, 商品名称, 货号, SKU货号, 品牌, 规格, 数量, 出价金额(元), ...
# Find column indices
header = {}
for c in range(1, ws.max_column + 1):
    h = ws.cell(1, c).value
    if h: header[str(h).strip()] = c

col_order_status = None
col_pay_time = None
col_order_time = None
for c in range(1, ws.max_column + 1):
    h = str(ws.cell(1, c).value or "").strip()
    if "订单状态" in h:
        col_order_status = c
    elif "支付时间" in h or "付款时间" in h:
        col_pay_time = c
    elif "订单创建时间" in h or "下单时间" in h:
        col_order_time = c

print(f"  order_status col={col_order_status}, pay_time={col_pay_time}, order_time={col_order_time}", file=sys.stderr)

spu_id_col = 3  # spuID
sku_id_col = 4  
name_col = 5
brand_col = 8
amount_col = 11
qty_col = 10

VALID_STATUSES = ["交易成功", "待发货", "待收货", "待买家收货", "待平台发货", "待卖家发货", "待平台收货"]

orders_monthly = defaultdict(lambda: defaultdict(lambda: {"gmv": 0.0, "orders": 0}))
orders_by_spu = defaultdict(lambda: defaultdict(lambda: {"gmv": 0.0, "orders": 0}))

processed = 0
for r in range(2, ws.max_row + 1):
    spuid = ws.cell(r, spu_id_col).value
    amount = ws.cell(r, amount_col).value
    brand = ws.cell(r, brand_col).value
    status = ws.cell(r, col_order_status).value if col_order_status else None
    
    # Get date
    date_str = None
    if col_pay_time:
        dt = ws.cell(r, col_pay_time).value
        date_str = excel_to_date(dt)
    if not date_str and col_order_time:
        dt = ws.cell(r, col_order_time).value
        date_str = excel_to_date(dt)
    
    if not (spuid and amount and brand in BRANDS and status in VALID_STATUSES and date_str and in_range(date_str)):
        continue
    
    b = BRAND_SHORT[brand]
    month = date_to_month(date_str)
    gmv = float(amount) * (ws.cell(r, qty_col).value or 1)
    
    orders_monthly[b][month]["gmv"] += gmv
    orders_monthly[b][month]["orders"] += 1
    
    spu_key = spu_name.get(int(spuid), str(spuid))
    orders_by_spu[b][spu_key]["gmv"] += gmv
    orders_by_spu[b][spu_key]["orders"] += 1
    orders_by_spu[b][spu_key]["spuid"] = int(spuid)
    
    processed += 1

print(f"  Processed {processed} valid orders", file=sys.stderr)

# Convert to sorted lists
orders_monthly_out = {}
for brand in BRAND_SHORT.values():
    orders_monthly_out[brand] = {}
    for m in sorted(orders_monthly[brand].keys()):
        orders_monthly_out[brand][m] = {
            "gmv": round(orders_monthly[brand][m]["gmv"], 2),
            "orders": orders_monthly[brand][m]["orders"]
        }

orders_by_spu_out = {}
for brand in BRAND_SHORT.values():
    spus = []
    for spu_name_key, data in sorted(orders_by_spu[brand].items(), key=lambda x: -x[1]["gmv"]):
        spus.append({
            "spu": spu_name_key,
            "spuid": data["spuid"],
            "gmv": round(data["gmv"], 2),
            "orders": data["orders"]
        })
    orders_by_spu_out[brand] = spus

# ============================================================
# 4. 商详访客数据 - UV按月/按SPU
# ============================================================
print("4. Extracting 商详访客数据...", file=sys.stderr)
ws = wb["商详访客数据"]
# Columns: 月份, 日期, SPUID, 商品名称, 货号, 品牌名称, 类目名称, 商品类型, 出价渠道, 支付订单金额, 支付订单量, ...
# Find the key columns
uv_header = {}
for c in range(1, min(50, ws.max_column + 1)):
    h = ws.cell(1, c).value
    if h: uv_header[str(h).strip()] = c

# Check what columns exist
print(f"  商详 headers (first 20 cols): {[ws.cell(1,c).value for c in range(1,21)]}", file=sys.stderr)

# Find indices: SPUID=col3, 品牌=col6, 商详访客人数 column
spuid_col_uv = 3
brand_col_uv = 6
# Look for 商详访客人数 column
uv_col = None
pay_amount_col = None
pay_orders_col = None
for c in range(1, min(50, ws.max_column + 1)):
    h = str(ws.cell(1, c).value or "").strip()
    # Specifically look for 商详访问指数（UV）or similar
    if "商详访问" in h and ("uv" in h.lower() or "指数" in h):
        if uv_col is None:
            uv_col = c
    elif "支付订单金额" in h:
        pay_amount_col = c
    elif "支付订单量" in h:
        pay_orders_col = c

print(f"  UV col={uv_col}, pay_amount={pay_amount_col}, pay_orders={pay_orders_col}", file=sys.stderr)

uv_monthly = defaultdict(lambda: defaultdict(lambda: {"uv": 0, "gmv": 0.0, "orders": 0}))
uv_by_spu = defaultdict(lambda: defaultdict(lambda: {"daily": [], "monthly": defaultdict(lambda: {"uv": 0, "gmv": 0.0, "orders": 0})}))

uv_processed = 0
for r in range(2, ws.max_row + 1):
    spuid = ws.cell(r, spuid_col_uv).value
    brand = ws.cell(r, brand_col_uv).value
    
    # Get date - column 2 is 日期
    date_cell = ws.cell(r, 2).value
    date_str = excel_to_date(date_cell)
    
    if not (spuid and brand in BRANDS and date_str and in_range(date_str)):
        continue
    
    b = BRAND_SHORT[brand]
    month = date_to_month(date_str)
    
    uv = ws.cell(r, uv_col).value if uv_col else 0
    pay_amt = ws.cell(r, pay_amount_col).value if pay_amount_col else 0
    pay_ord = ws.cell(r, pay_orders_col).value if pay_orders_col else 0
    
    def clean_num(v):
        """Extract number from values like '122 (指数)' or '0 (指数)'"""
        if v is None:
            return 0
        if isinstance(v, (int, float)):
            return v
        if isinstance(v, str):
            # Take only the part before space/parenthesis
            import re
            m = re.match(r'[\d.]+', v.strip())
            return float(m.group()) if m else 0
        return 0
    
    uv_val = int(clean_num(uv))
    pay_amt_val = float(clean_num(pay_amt))
    pay_ord_val = int(clean_num(pay_ord))
    
    uv_monthly[b][month]["uv"] += uv_val
    uv_monthly[b][month]["gmv"] += pay_amt_val
    uv_monthly[b][month]["orders"] += pay_ord_val
    
    spu_key = spu_name.get(int(spuid), str(spuid))
    uv_by_spu[b][spu_key]["daily"].append({
        "date": date_str,
        "uv": uv_val,
        "gmv": round(pay_amt_val, 2),
        "orders": pay_ord_val
    })
    uv_by_spu[b][spu_key]["monthly"][month]["uv"] += uv_val
    uv_by_spu[b][spu_key]["monthly"][month]["gmv"] += pay_amt_val
    uv_by_spu[b][spu_key]["monthly"][month]["orders"] += pay_ord_val
    
    uv_processed += 1

print(f"  UV processed {uv_processed} records", file=sys.stderr)

uv_monthly_out = {}
for brand in BRAND_SHORT.values():
    uv_monthly_out[brand] = {}
    for m in sorted(uv_monthly[brand].keys()):
        uv_monthly_out[brand][m] = {
            "uv": uv_monthly[brand][m]["uv"],
            "gmv": round(uv_monthly[brand][m]["gmv"], 2),
            "orders": uv_monthly[brand][m]["orders"]
        }

uv_by_spu_out = {}
for brand in BRAND_SHORT.values():
    uv_by_spu_out[brand] = {}
    for spu_key, data in uv_by_spu[brand].items():
        uv_by_spu_out[brand][spu_key] = {
            "daily": sorted(data["daily"], key=lambda x: x["date"]),
            "monthly": {m: {"uv": d["uv"], "gmv": round(d["gmv"], 2), "orders": d["orders"]} 
                       for m, d in sorted(data["monthly"].items())}
        }

# ============================================================
# 5. 得物推数据 - 按月汇总（双源品牌匹配：货盘表 + 交易订单）
# ============================================================
print("5. Extracting 得物推数据...", file=sys.stderr)
ws = wb["得物推数据"]
# Headers: 时间, 用户ID, 计划名称, 计划ID, 计划类型, 优化目标, 商品ID, 消耗(元), 曝光(次), 点击(次), 点击率(%)...
push_header = {}
for c in range(1, min(30, ws.max_column + 1)):
    h = ws.cell(1, c).value
    if h: push_header[str(h).strip()] = c

time_col = 1
goods_id_col = 7
cost_col = 8
print(f"  Push: time_col={time_col}, goods_id_col={goods_id_col}, cost_col={cost_col}", file=sys.stderr)

# Build SPU->brand mapping from 交易订单 too (supplement 货盘表)
spu_brand_from_orders = {}
ws_orders = wb["交易订单"]
spu_id_col_ord = 3
brand_col_ord = 8
for r in range(2, ws_orders.max_row + 1):
    spuid = ws_orders.cell(r, spu_id_col_ord).value
    brand = ws_orders.cell(r, brand_col_ord).value
    if spuid and brand in BRANDS:
        try:
            spu_brand_from_orders[int(spuid)] = BRAND_SHORT[brand]
        except (ValueError, TypeError):
            pass
print(f"  Orders SPU→brand mapping: {len(spu_brand_from_orders)} SPUs", file=sys.stderr)

push_monthly = defaultdict(lambda: defaultdict(float))
push_processed = 0
for r in range(2, ws.max_row + 1):
    dt = ws.cell(r, time_col).value
    date_str = excel_to_date(dt)
    good_id = ws.cell(r, goods_id_col).value
    cost = ws.cell(r, cost_col).value
    
    if not (date_str and good_id and cost and in_range(date_str)):
        continue
    
    # Match brand: first from 货盘表, then fallback to 交易订单
    brand = None
    try:
        gid = int(good_id)
        brand = spu_brand.get(gid)
        if not brand:
            brand = spu_brand_from_orders.get(gid)
    except (ValueError, TypeError):
        continue
    
    if not brand:
        continue
    
    month = date_to_month(date_str)
    push_monthly[brand][month] += float(cost)
    push_processed += 1

print(f"  Push processed {push_processed} records", file=sys.stderr)

push_monthly_out = {}
for brand in BRAND_SHORT.values():
    push_monthly_out[brand] = {m: round(v, 2) for m, v in sorted(push_monthly[brand].items())}

# ============================================================
# 6. 社区投放数据 - 按月汇总
# ============================================================
print("6. Extracting 社区投放数据...", file=sys.stderr)
ws = wb["社区投放数据"]
# Headers: 主任务ID, 子任务ID, 货号(spu_id), 任务类型, 达人昵称, 达人uid, 子任务金额, 动态发布时间...
comm_header = {}
for c in range(1, min(30, ws.max_column + 1)):
    h = ws.cell(1, c).value
    if h: comm_header[str(h).strip()] = c

spu_id_col_comm = 3  # 货号(spu_id)
amount_col_comm = 7   # 子任务金额
time_col_comm = 8     # 动态发布时间

comm_monthly = defaultdict(lambda: defaultdict(lambda: {"cost": 0.0, "tasks": 0}))
comm_processed = 0
for r in range(2, ws.max_row + 1):
    spu_ids = ws.cell(r, spu_id_col_comm).value
    amount_str = ws.cell(r, amount_col_comm).value
    dt = ws.cell(r, time_col_comm).value
    date_str = excel_to_date(dt)
    
    if not (spu_ids and amount_str and date_str and in_range(date_str)):
        continue
    
    # Parse amount (could be "¥200元" or 200)
    try:
        if isinstance(amount_str, str):
            amount_str = amount_str.replace("¥", "").replace("元", "").replace(",", "").strip()
        amount = float(amount_str)
    except:
        continue
    
    # Parse SPU IDs (could be "29249125,25213007,...")
    spu_list = str(spu_ids).split(",")
    
    month = date_to_month(date_str)
    brands_for_task = set()
    for sid in spu_list:
        try:
            b = spu_brand.get(int(sid.strip()))
            if b:
                brands_for_task.add(b)
        except:
            pass
    
    if not brands_for_task:
        continue
    
    # Split cost evenly among brands if multiple brands in one task
    per_brand = amount / len(brands_for_task)
    for b in brands_for_task:
        comm_monthly[b][month]["cost"] += per_brand
        comm_monthly[b][month]["tasks"] += 1
    
    comm_processed += 1

print(f"  Comm processed {comm_processed} records", file=sys.stderr)

comm_monthly_out = {}
for brand in BRAND_SHORT.values():
    comm_monthly_out[brand] = {}
    for m in sorted(comm_monthly[brand].keys()):
        comm_monthly_out[brand][m] = {
            "cost": round(comm_monthly[brand][m]["cost"], 2),
            "tasks": comm_monthly[brand][m]["tasks"]
        }

# ============================================================
# 7. Build output
# ============================================================
months_list = sorted(set(
    list(orders_monthly_out.get("卡西欧", {}).keys()) +
    list(orders_monthly_out.get("蔻驰", {}).keys()) +
    list(market_monthly_avg.keys())
))

output = {
    "dateRange": {"start": START_DATE, "end": END_DATE},
    "months": months_list,
    "orders_monthly": orders_monthly_out,
    "orders_by_spu": orders_by_spu_out,
    "uv_monthly": uv_monthly_out,
    "uv_by_spu": uv_by_spu_out,
    "market": market_data,
    "market_monthly_avg": market_monthly_avg,
    "market_bag": market_data_bag,
    "market_monthly_bag_avg": market_monthly_bag_avg,
    "push_monthly": push_monthly_out,
    "comm_monthly": comm_monthly_out
}

# Write as data.js
OUTPUT_PATH = "/home/Vic/dewu-reports/喜过/2026-01-05/data.js"
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    f.write("var D=")
    json.dump(output, f, ensure_ascii=False, separators=(",", ":"))
    f.write(";\n")

print(f"\n✅ Written to {OUTPUT_PATH}", file=sys.stderr)
print(f"  Size: {len(json.dumps(output, ensure_ascii=False))} chars", file=sys.stderr)
print(f"  Market records: {len(market_data)}", file=sys.stderr)
print(f"  Months: {months_list}", file=sys.stderr)
